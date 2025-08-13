import os
import glob
from openpyxl import load_workbook
from datetime import datetime
from tkinter import Tk, filedialog
 
# return plant name based on first two letters of the string
def interpret_prefix(cell_value):
    if not isinstance(cell_value, str) or len(cell_value) < 2:
        return "Invalid"
   
    prefix = cell_value[:2].upper()
    mapping = {
        "AL": "Albert Lea",
        "DA": "Danville",
        "MC": "Mason City",
        "NE": "Newton",
        "NT": "Newton",
        "SE": "Seneca",
        "EM": "Emden",
        "OE": "Oeding"
    }
    return mapping.get(prefix, "Unknown")
 
def main(folder_path = os.getcwd(), destination_file= os.getcwd()):
    # Define folder path and destination file
    # folder_path = input(r"C:\Users\vqqs\OneDrive - Chevron\Desktop\Automate\test_folder")
    # destination_file = input(r"C:\Users\vqqs\OneDrive - Chevron\Desktop\Automate\MWM Scheduling Data.xlsx")
    wb_dest = load_workbook(destination_file)
    ws_dest = wb_dest["Scheduled WOs"]
   
    # Loop through all .xlsm files in the folder
    for source_file in glob.glob(os.path.join(folder_path, "*.xlsm")):
        try:
            wb = load_workbook(source_file, data_only=True)
            ws_source = wb["Import WO Report"]
            ws_metrics = wb["Metrics Data"]
            date_location = wb["Man Hours"]
            prefix_cell_value = ws_source["H13"].value #Tag number of first work order in Import WO Report Sheet
            plant_name = interpret_prefix(prefix_cell_value) #Use tag number of first work order to identify plant name
            monday_date = date_location["A3"].value #Date of WO
       
            # Find the last non empty row in column A
            next_row = 2  # Assume row 1 is headers
            while ws_dest.cell(row=next_row, column=1).value is not None:
                next_row += 1
   
            # Collect and add valid work order numbers
            for row in ws_source.iter_rows(min_row=2, max_col=1, values_only=True):
                wo_number = row[0]
                if isinstance(wo_number, (int, float)) and wo_number != 0:
                    ws_dest.cell(row=next_row, column=1, value=plant_name) #plant name
                    ws_dest.cell(row=next_row, column=2, value=wo_number) # wo number
                    cell = ws_dest.cell(row=next_row, column=3, value=monday_date) #date
                    cell.number_format = "mm/dd/yyyy"
                    next_row += 1
       
            # Copy over Labor Hours
            ws_labor = wb_dest["Labor Hours"]
   
            # Find the last non-empty row in Labor Hours sheet
            labor_next_row = 2  # Assuming row 1 is headers
            while ws_labor.cell(row=labor_next_row, column=1).value is not None:
                labor_next_row += 1
   
            for row in ws_metrics.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True):
                if all(cell is None for cell in row):
                    break  # Stop copying if the entire row is empty
                ws_labor.cell(row=labor_next_row, column=1, value=plant_name)  # Column A: plant name
                for col_index, value in enumerate(row, start=2):  # Columns B to F
                    cell = ws_labor.cell(row=labor_next_row, column=col_index, value=value)
                    if col_index == 2 and isinstance(value, datetime):
                        cell.value = value.date()  # Strip time portion
                        cell.number_format = 'mm/dd/yyyy'
                labor_next_row += 1
   
            print(f"Processed: {os.path.basename(source_file)}")
   
        except Exception as e:
            print(f"Error processing {source_file}: {e}")
   
    # Save changes
    wb_dest.save(destination_file)
    print("All data appended successfully.")
 
if __name__ == "__main__":
 
    # Hide the root window
    root = Tk()
    root.withdraw()
 
    # Select a folder
    folder_path = filedialog.askdirectory()
    print("Selected folder:", folder_path)
 
    # Or select a file
    destination_path = filedialog.askopenfilename()
    print("Selected file:", destination_path)
 
    main(folder_path, destination_path)